"""
MediaHub — Медиахаб для молодёжных центров
FastAPI + PostgreSQL backend
"""

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List
import psycopg2
import psycopg2.extras
import json, os, random, uuid, shutil, hashlib, io
from urllib.parse import quote
import requests as http_requests
from datetime import datetime, timedelta
from dotenv import load_dotenv
from jose import jwt, JWTError
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

app = FastAPI(title="MediaHub API", version="1.0.0")

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_ORIGINS = [
    o.strip().rstrip("/")
    for o in os.getenv(
        "ALLOWED_ORIGINS",
        "http://localhost:3000,http://127.0.0.1:3000,https://frontend-production-cd62.up.railway.app",
    ).split(",")
    if o.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://mediahub:mediahub123@localhost:5432/mediahub")

# ── Password helpers ─────────────────────────────────────────────────────────

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed

# ── JWT helpers ──────────────────────────────────────────────────────────────

JWT_SECRET = os.getenv("JWT_SECRET", "changeme-in-production-2025")
JWT_ALGORITHM = "HS256"

def create_token(user_id: int) -> str:
    expire = datetime.utcnow() + timedelta(hours=72)
    return jwt.encode({"sub": str(user_id), "exp": expire}, JWT_SECRET, algorithm=JWT_ALGORITHM)

def get_current_user_id(authorization: str = Header(None)) -> int:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Требуется авторизация")
    try:
        token = authorization.split(" ", 1)[1]
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return int(payload["sub"])
    except (JWTError, KeyError, ValueError, IndexError):
        raise HTTPException(401, "Недействительный токен")

def require_group_member(group_id: int, user_id: int, conn) -> str:
    c = conn.cursor()
    c.execute("SELECT role FROM group_members WHERE group_id=%s AND user_id=%s", (group_id, user_id))
    row = c.fetchone()
    if not row:
        raise HTTPException(403, "Нет доступа к этой группе")
    return row["role"]

# ── Pydantic models ──────────────────────────────────────────────────────────

class MediaItem(BaseModel):
    url: str
    type: str      # "image" | "video"
    filename: str

class PostCreate(BaseModel):
    title: str
    content: str
    status: str = "draft"
    platforms: List[str] = ["vk"]
    tags: List[str] = []
    scheduled_at: Optional[str] = None
    template_type: Optional[str] = None
    author_id: int = 1
    media: List[MediaItem] = []
    location_address: Optional[str] = None
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None

class PostUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    status: Optional[str] = None
    platforms: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    scheduled_at: Optional[str] = None
    media: Optional[List[MediaItem]] = None
    location_address: Optional[str] = None
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None

class GenerateRequest(BaseModel):
    template_type: str
    fields: dict

class UserUpdate(BaseModel):
    role: str

class LoginRequest(BaseModel):
    email: str
    password: str

class AIEnhanceRequest(BaseModel):
    text: str
    mode: str  # 'creative' | 'russify'

class VkSettingsSave(BaseModel):
    group_id: str
    access_token: str

class TgSettingsSave(BaseModel):
    bot_token: str
    chat_id: str

class VkOAuthExchange(BaseModel):
    app_id: str
    app_secret: str
    code: str
    group_id: str

class RegisterRequest(BaseModel):
    name: str
    email: str
    password: str

class UserCreate(BaseModel):
    name: str
    email: str
    role: str = "editor"
    password: str

class GroupCreate(BaseModel):
    name: str
    description: str = ""

class GroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    avatar: Optional[str] = None

class GroupMemberRoleUpdate(BaseModel):
    role: str

class InviteLinkCreate(BaseModel):
    role: str = "editor"
    expires_hours: int = 24
    max_uses: Optional[int] = None

# ── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)

def row_to_dict(row):
    if row is None:
        return None
    d = dict(row)
    d.pop("password_hash", None)
    for key in ("platforms", "tags", "media"):
        if key in d and isinstance(d[key], str):
            try:
                d[key] = json.loads(d[key])
            except Exception:
                d[key] = []
        elif key not in d:
            d[key] = []
    return d

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            role TEXT NOT NULL DEFAULT 'editor',
            avatar TEXT DEFAULT '',
            created_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS posts (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            content TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft',
            platforms TEXT DEFAULT '["vk"]',
            tags TEXT DEFAULT '[]',
            scheduled_at TEXT,
            published_at TEXT,
            views INTEGER DEFAULT 0,
            reactions INTEGER DEFAULT 0,
            comments INTEGER DEFAULT 0,
            shares INTEGER DEFAULT 0,
            author_id INTEGER DEFAULT 1,
            template_type TEXT,
            created_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI'),
            FOREIGN KEY (author_id) REFERENCES users(id)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS templates (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT NOT NULL UNIQUE,
            description TEXT,
            fields TEXT DEFAULT '[]',
            template_text TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,
            user_id INTEGER DEFAULT 1,
            message TEXT NOT NULL,
            type TEXT DEFAULT 'info',
            is_read INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS vk_settings (
            id INTEGER PRIMARY KEY DEFAULT 1,
            group_id TEXT NOT NULL,
            access_token TEXT NOT NULL,
            group_name TEXT DEFAULT '',
            connected_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS tg_settings (
            id INTEGER PRIMARY KEY DEFAULT 1,
            bot_token TEXT NOT NULL,
            chat_id TEXT NOT NULL,
            chat_title TEXT DEFAULT '',
            connected_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    c.execute("ALTER TABLE posts ADD COLUMN IF NOT EXISTS media TEXT DEFAULT '[]'")
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS password_hash TEXT")
    c.execute("ALTER TABLE posts ADD COLUMN IF NOT EXISTS location_address TEXT")
    c.execute("ALTER TABLE posts ADD COLUMN IF NOT EXISTS location_lat DOUBLE PRECISION")
    c.execute("ALTER TABLE posts ADD COLUMN IF NOT EXISTS location_lng DOUBLE PRECISION")

    # ── Groups and memberships ───────────────────────────────────────────────
    c.execute("""
        CREATE TABLE IF NOT EXISTS groups (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            avatar TEXT DEFAULT '',
            created_by INTEGER NOT NULL REFERENCES users(id),
            created_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS group_members (
            id SERIAL PRIMARY KEY,
            group_id INTEGER NOT NULL REFERENCES groups(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            role TEXT NOT NULL DEFAULT 'editor',
            joined_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI'),
            UNIQUE (group_id, user_id)
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS invite_links (
            id SERIAL PRIMARY KEY,
            group_id INTEGER NOT NULL REFERENCES groups(id) ON DELETE CASCADE,
            token TEXT NOT NULL UNIQUE,
            role TEXT NOT NULL DEFAULT 'editor',
            created_by INTEGER NOT NULL REFERENCES users(id),
            expires_at TEXT NOT NULL,
            used_count INTEGER DEFAULT 0,
            max_uses INTEGER DEFAULT 1,
            created_at TEXT DEFAULT to_char(CURRENT_TIMESTAMP, 'YYYY-MM-DD"T"HH24:MI')
        )
    """)

    # ── Add group_id to existing tables ──────────────────────────────────────
    c.execute("ALTER TABLE posts ADD COLUMN IF NOT EXISTS group_id INTEGER REFERENCES groups(id)")
    c.execute("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS group_id INTEGER REFERENCES groups(id)")
    # workspace_id links vk/tg rows to app groups (group_id TEXT is already taken by VK group id)
    c.execute("ALTER TABLE vk_settings ADD COLUMN IF NOT EXISTS workspace_id INTEGER REFERENCES groups(id)")
    c.execute("ALTER TABLE tg_settings ADD COLUMN IF NOT EXISTS workspace_id INTEGER REFERENCES groups(id)")

    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()["count"] == 0:
        c.executemany(
            "INSERT INTO users (name, email, role, avatar) VALUES (%s, %s, %s, %s)",
            [
                ("Алексей Иванов",  "admin@mediahub.ru",    "admin",    "АИ"),
                ("Мария Петрова",   "editor@mediahub.ru",   "editor",   "МП"),
                ("Дмитрий Сидоров", "observer@mediahub.ru", "observer", "ДС"),
            ],
        )

    c.execute("SELECT COUNT(*) FROM templates")
    if c.fetchone()["count"] == 0:
        c.executemany(
            "INSERT INTO templates (name, type, description, fields, template_text) VALUES (%s, %s, %s, %s, %s)",
            [
                ("Анонс мероприятия", "announcement", "Объявление о предстоящем событии",
                 json.dumps([
                     {"key": "event_name",   "label": "Название мероприятия", "placeholder": "Хакатон «IT-Кубок»"},
                     {"key": "date",         "label": "Дата и время",         "placeholder": "25 апреля, 10:00"},
                     {"key": "location",     "label": "Место проведения",     "placeholder": "МЦ «Зеркало»"},
                     {"key": "description",  "label": "Краткое описание",     "placeholder": "Соревнования по программированию"},
                     {"key": "contact",      "label": "Контакт для записи",   "placeholder": "@smm_manager"},
                 ]),
                 "🔥 {event_name}\n\n📅 Дата: {date}\n📍 Место: {location}\n\n{description}\n\n👉 Успей зарегистрироваться! Контакт: {contact}\n\n#мероприятие #молодёжь #красноярск"),
                ("Итоги события", "results", "Публикация результатов прошедшего мероприятия",
                 json.dumps([
                     {"key": "event_name",    "label": "Название мероприятия",   "placeholder": "Форум молодых лидеров"},
                     {"key": "participants",  "label": "Кол-во участников",       "placeholder": "150"},
                     {"key": "highlights",    "label": "Главные моменты",         "placeholder": "5 спикеров, мастер-классы"},
                     {"key": "next_event",    "label": "Следующее мероприятие",   "placeholder": "Следующий форум — в июне"},
                 ]),
                 "✅ {event_name} — позади!\n\n👥 Участников: {participants}\n\n🎯 {highlights}\n\nСпасибо всем! {next_event} — следите за анонсами.\n\n#итоги #молодёжь #красноярск"),
                ("Вакансия", "vacancy", "Объявление об открытой позиции",
                 json.dumps([
                     {"key": "position",     "label": "Должность",    "placeholder": "SMM-менеджер"},
                     {"key": "organization", "label": "Организация",  "placeholder": "МЦ «Зеркало»"},
                     {"key": "requirements", "label": "Требования",   "placeholder": "Опыт от 1 года"},
                     {"key": "conditions",   "label": "Условия",      "placeholder": "Гибкий график"},
                     {"key": "contact",      "label": "Контакт",      "placeholder": "@hr_manager"},
                 ]),
                 "🚀 Вакансия: {position}\n🏢 {organization}\n\n📋 Требования:\n{requirements}\n\n💼 Условия:\n{conditions}\n\n📩 Откликнуться: {contact}\n\n#вакансия #работа #красноярск"),
                ("Грант", "grant", "Информация о грантовой программе",
                 json.dumps([
                     {"key": "grant_name", "label": "Название гранта",  "placeholder": "Грант «Молодёжь края»"},
                     {"key": "amount",     "label": "Размер поддержки", "placeholder": "до 500 000 ₽"},
                     {"key": "deadline",   "label": "Дедлайн подачи",   "placeholder": "1 мая 2026"},
                     {"key": "who",        "label": "Для кого",         "placeholder": "НКО и молодёжные орг."},
                     {"key": "link",       "label": "Ссылка",           "placeholder": "grant.krasn.ru"},
                 ]),
                 "💰 {grant_name}\n\n🎁 Поддержка: {amount}\n⏰ Дедлайн: {deadline}\n\n👥 Для кого: {who}\n\n🔗 Подробнее: {link}\n\n#грант #поддержка #молодёжь"),
            ],
        )

    c.execute("SELECT COUNT(*) FROM posts")
    if c.fetchone()["count"] == 0:
        now = datetime.now()
        rows = [
            ("🔥 Хакатон IT-Кубок — регистрация открыта!",
             "🔥 Хакатон IT-Кубок\n\n📅 Дата: 25–26 апреля, 10:00\n📍 Место: Каменка\n\nСоревнования по программированию.\n\n👉 @it_kubok\n\n#мероприятие #молодёжь",
             "published", ["vk", "telegram"], ["мероприятия"], -7, 1850, 142, 38, 24, "announcement"),
            ("✅ Форум молодых лидеров — итоги!",
             "✅ Форум молодых лидеров — позади!\n\n👥 Участников: 200\n\n🎯 7 спикеров, 4 мастер-класса.\n\n#итоги #молодёжь",
             "published", ["vk"], ["мероприятия"], -14, 2100, 198, 55, 41, "results"),
            ("🚀 Вакансия: SMM-менеджер в МЦ «Зеркало»",
             "🚀 SMM-менеджер\n🏢 МЦ «Зеркало»\n\n📋 Опыт от 1 года\n💼 Гибкий график\n📩 @hr_zerkalo\n\n#вакансия",
             "published", ["vk", "telegram"], ["вакансии"], -5, 980, 64, 12, 8, "vacancy"),
            ("💰 Грант «Молодёжь края» — подай заявку",
             "💰 Грант «Молодёжь края»\n\n🎁 до 500 000 ₽\n⏰ 1 мая 2026\n🔗 grant.krasn.ru\n\n#грант",
             "published", ["vk"], ["гранты"], -3, 1340, 87, 21, 15, "grant"),
            ("🌟 Волонтёрская смена — набор участников",
             "🌟 Волонтёрская смена 2026\n\n📅 1–7 июня\n📍 Столбы\n\n#волонтёр",
             "scheduled", ["vk", "telegram"], ["мероприятия"], 2, 0, 0, 0, 0, "announcement"),
            ("📊 Итоги апреля — наша статистика",
             "📊 Апрель:\n\n✅ 24 публикации\n👥 Охват: 1 200\n❤️ Вовлечённость: 8.3%\n\n#статистика",
             "scheduled", ["vk"], ["новости"], 1, 0, 0, 0, 0, None),
            ("🎓 Интенсив «Стартап за 48 часов»",
             "🎓 Интенсив «Стартап за 48 часов»\n\n📅 10 мая\n📍 Краевой дворец молодёжи\n\n#стартап",
             "draft", ["vk", "telegram"], ["мероприятия"], 5, 0, 0, 0, 0, "announcement"),
            ("🏋️ Спортивный фестиваль «Активная молодёжь»",
             "🏋️ «Активная молодёжь»\n\n📅 18 мая\n📍 Стадион «Рассвет»\n\n#спорт",
             "draft", ["vk"], ["мероприятия"], 10, 0, 0, 0, 0, None),
        ]
        for title, content, status, platforms, tags, days, views, reactions, comments, shares, tmpl in rows:
            dt = now + timedelta(days=days)
            scheduled_at = dt.strftime("%Y-%m-%dT%H:%M") if status == "scheduled" else None
            published_at  = dt.strftime("%Y-%m-%dT%H:%M") if status == "published"  else None
            created_at    = (now + timedelta(days=days - 1)).strftime("%Y-%m-%dT%H:%M")
            c.execute(
                "INSERT INTO posts (title,content,status,platforms,tags,scheduled_at,published_at,views,reactions,comments,shares,author_id,template_type,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (title, content, status, json.dumps(platforms), json.dumps(tags), scheduled_at, published_at, views, reactions, comments, shares, random.choice([1, 2]), tmpl, created_at),
            )

        c.executemany(
            "INSERT INTO notifications (user_id, message, type, is_read) VALUES (%s, %s, %s, %s)",
            [
                (1, "Пост «Хакатон IT-Кубок» опубликован",          "success", 0),
                (1, "Запланирован пост на завтра: «Итоги апреля»",  "info",    0),
                (1, "Черновик «Спортивный фестиваль» не опубликован 5 дней", "warning", 1),
                (2, "Новый пост на модерации от волонтёра",          "info",    0),
            ],
        )

    # ── Default group migration ──────────────────────────────────────────────────
    c.execute("SELECT COUNT(*) FROM groups")
    if c.fetchone()["count"] == 0:
        c.execute("SELECT id FROM users ORDER BY id ASC LIMIT 1")
        first_user_row = c.fetchone()
        if first_user_row:
            first_user_id = first_user_row["id"]
            c.execute(
                "INSERT INTO groups (name, description, created_by) VALUES (%s, %s, %s) RETURNING id",
                ("Медиа-Хаб", "Группа по умолчанию", first_user_id)
            )
            gid = c.fetchone()["id"]

            c.execute("SELECT id FROM users")
            for u in c.fetchall():
                c.execute(
                    "INSERT INTO group_members (group_id, user_id, role) VALUES (%s, %s, 'admin') ON CONFLICT DO NOTHING",
                    (gid, u["id"])
                )

            c.execute("UPDATE posts SET group_id=%s WHERE group_id IS NULL", (gid,))
            c.execute("UPDATE notifications SET group_id=%s WHERE group_id IS NULL", (gid,))
            c.execute("UPDATE vk_settings SET workspace_id=%s WHERE workspace_id IS NULL", (gid,))
            c.execute("UPDATE tg_settings SET workspace_id=%s WHERE workspace_id IS NULL", (gid,))

    conn.commit()
    conn.close()

# ── Health ───────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "service": "MediaHub API"}

# ── File upload ──────────────────────────────────────────────────────────────

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
ALLOWED_VIDEO_TYPES = {"video/mp4", "video/quicktime", "video/webm", "video/x-msvideo"}
ALLOWED_DOC_TYPES = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-powerpoint",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "text/plain",
    "text/csv",
}
MAX_IMAGE_SIZE = 10 * 1024 * 1024   # 10 MB
MAX_VIDEO_SIZE = 100 * 1024 * 1024  # 100 MB
MAX_DOC_SIZE = 50 * 1024 * 1024     # 50 MB (VK doc limit is 200 MB)

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    content_type = file.content_type or ""
    if content_type not in ALLOWED_IMAGE_TYPES | ALLOWED_VIDEO_TYPES | ALLOWED_DOC_TYPES:
        raise HTTPException(400, f"Неподдерживаемый тип файла: {content_type}")

    ext = (file.filename or "file").rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "bin"
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    data = await file.read()
    if content_type in ALLOWED_VIDEO_TYPES:
        max_size = MAX_VIDEO_SIZE
    elif content_type in ALLOWED_DOC_TYPES:
        max_size = MAX_DOC_SIZE
    else:
        max_size = MAX_IMAGE_SIZE
    if len(data) > max_size:
        raise HTTPException(400, f"Файл слишком большой (макс. {max_size // 1024 // 1024} МБ)")

    with open(filepath, "wb") as f:
        f.write(data)

    if content_type in ALLOWED_VIDEO_TYPES:
        file_type = "video"
    elif content_type in ALLOWED_DOC_TYPES:
        file_type = "doc"
    else:
        file_type = "image"
    return {"url": f"/uploads/{filename}", "type": file_type, "filename": file.filename or filename}

# ── AI Enhance ───────────────────────────────────────────────────────────────

_AI_PROMPTS = {
    "creative": (
        "Ты — опытный SMM-редактор молодёжного центра. Перепиши текст так, чтобы он "
        "цеплял аудиторию 16–30 лет с первой строки: замени канцеляризмы живыми словами, "
        "добавь энергию и искренние эмоции, сделай ритм лёгким и читаемым. "
        "Вставь эмодзи там, где они усиливают смысл или настроение — не переусердствуй, "
        "1–3 эмодзи в нужных местах лучше, чем россыпь везде. "
        "Сохрани все факты, даты, имена и длину оригинала. "
        "Верни только готовый текст, без комментариев и пояснений."
    ),
    "formal": (
        "Ты — пресс-секретарь молодёжного центра с опытом работы в госструктурах. "
        "Перепиши текст в официальном, но живом стиле: грамотно, структурированно, "
        "без сленга и излишней эмоциональности, но и без бюрократической сухости. "
        "Используй чёткие формулировки, активный залог, уважительный тон. "
        "Текст должен подходить для официальных анонсов, партнёрских постов и отчётов. "
        "Сохрани все факты и структуру. Верни только готовый текст, без пояснений."
    ),
    "calltoaction": (
        "Ты — копирайтер молодёжного центра. Оставь основной текст без изменений "
        "и добавь в конец яркий, мотивирующий призыв к действию для аудитории 16–30 лет. "
        "Выбери глагол по смыслу поста: записаться, прийти, написать нам, подать заявку, "
        "узнать подробнее, поделиться с друзьями — и т.д. "
        "Тон — дружеский и воодушевляющий, без давления и манипуляций. "
        "Добавь 1–2 уместных эмодзи в призыв, чтобы он выделялся визуально. "
        "Верни полный текст с добавленным призывом, без комментариев."
    ),
    "shortify": (
        "Ты — редактор с острым чувством слова. Сократи текст примерно вдвое: "
        "безжалостно убери воду, повторы, лишние вводные слова и затянутые конструкции. "
        "Сохрани главную мысль, все ключевые факты (даты, имена, цифры) и живой тон. "
        "Не добавляй ничего нового. Верни только сокращённый вариант, без пояснений."
    ),
    "hashtags": (
        "Ты — SMM-специалист молодёжного центра. Проанализируй тему поста и придумай "
        "5–7 релевантных хештегов: микс из широких (#молодёжь, #события) и нишевых "
        "(по конкретной теме поста). Часть хештегов — на русском, часть — на английском. "
        "Хештеги должны реально использоваться в ВКонтакте и Telegram. "
        "Верни только хештеги через пробел, без текста поста и без пояснений."
    ),
    "russify": (
        "Ты — редактор русского языка. Пройдись по тексту и замени все англицизмы, "
        "заимствованный сленг и кальки на естественные русские аналоги, которые не режут слух. "
        "Примеры: контент → материал, дедлайн → срок, фидбек → отклик, "
        "постить → публиковать, ивент → мероприятие, воркшоп → мастер-класс. "
        "Сохрани стиль, тон и структуру текста. Имена, названия и аббревиатуры не трогай. "
        "Верни только исправленный текст, без пояснений и списка замен."
    ),
}

@app.post("/api/ai-enhance")
def ai_enhance(body: AIEnhanceRequest):
    if not body.text.strip():
        raise HTTPException(400, "Текст не может быть пустым")

    system_prompt = _AI_PROMPTS.get(body.mode)
    if not system_prompt:
        raise HTTPException(400, "Неверный режим. Допустимые значения: creative, russify, shortify, formal, hashtags, calltoaction")

    groq_key = os.getenv("GROQ_API_KEY", "")
    if not groq_key:
        raise HTTPException(503, (
            "GROQ_API_KEY не настроен. "
            "Получите бесплатный ключ на console.groq.com и добавьте его в .env"
        ))

    try:
        resp = http_requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
            json={
                "model": "llama-3.1-8b-instant",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": body.text},
                ],
                "temperature": {"creative": 0.75, "shortify": 0.3, "formal": 0.3,
                                "hashtags": 0.5, "calltoaction": 0.6, "russify": 0.25}.get(body.mode, 0.5),
                "max_tokens": 2000,
            },
            timeout=30,
        )
        resp.raise_for_status()
        result = resp.json()
        enhanced = result["choices"][0]["message"]["content"].strip()
        return {"text": enhanced}
    except http_requests.exceptions.Timeout:
        raise HTTPException(504, "Превышено время ожидания ответа от ИИ (30 с)")
    except http_requests.exceptions.HTTPError as e:
        detail = ""
        try:
            detail = e.response.json().get("error", {}).get("message", "")
        except Exception:
            pass
        raise HTTPException(502, f"Ошибка Groq API: {detail or str(e)}")
    except http_requests.exceptions.RequestException as e:
        raise HTTPException(502, f"Ошибка связи с ИИ: {str(e)}")
    except (KeyError, IndexError):
        raise HTTPException(502, "Неожиданный формат ответа от ИИ")

# ── VK API helpers ───────────────────────────────────────────────────────────

VK_API_VERSION = "5.199"

def vk_get_group_name(access_token: str, group_id: str) -> str:
    clean_id = group_id.lstrip("-")
    try:
        r = http_requests.get(
            "https://api.vk.com/method/groups.getById",
            params={"group_ids": clean_id, "access_token": access_token, "v": VK_API_VERSION},
            timeout=10,
        )
        data = r.json()
        if "error" in data:
            raise ValueError(data["error"].get("error_msg", "VK API error"))
        response = data.get("response", {})
        # API v5.196+ returns {"groups": [...]}; older versions return [...]
        if isinstance(response, list):
            groups = response
        else:
            groups = response.get("groups", [])
        if not groups:
            raise ValueError("Группа не найдена")
        return groups[0].get("name", "")
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Ошибка связи с VK: {e}")

def vk_upload_photo_to_wall(access_token: str, group_id: str, image_data: bytes, filename: str = "photo.jpg") -> str:
    clean_id = group_id.lstrip("-")
    r = http_requests.get(
        "https://api.vk.com/method/photos.getWallUploadServer",
        params={"group_id": clean_id, "access_token": access_token, "v": VK_API_VERSION},
        timeout=10,
    )
    data = r.json()
    if "error" in data:
        raise ValueError(data["error"].get("error_msg", "VK getWallUploadServer error"))
    upload_url = data["response"]["upload_url"]

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    mime_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "gif": "image/gif", "webp": "image/webp"}
    content_type = mime_map.get(ext, "image/jpeg")

    r2 = http_requests.post(upload_url, files={"photo": (filename, image_data, content_type)}, timeout=60)
    r2.raise_for_status()
    upload_result = r2.json()
    if "error" in upload_result:
        err = upload_result["error"]
        raise ValueError(err if isinstance(err, str) else str(err))
    if not upload_result.get("photo") or not upload_result.get("server"):
        raise ValueError(f"Неожиданный ответ от VK при загрузке фото: {upload_result}")

    r3 = http_requests.post(
        "https://api.vk.com/method/photos.saveWallPhoto",
        data={
            "group_id": clean_id,
            "photo": upload_result["photo"],
            "server": upload_result["server"],
            "hash": upload_result["hash"],
            "access_token": access_token,
            "v": VK_API_VERSION,
        },
        timeout=15,
    )
    saved = r3.json()
    if "error" in saved:
        raise ValueError(saved["error"].get("error_msg", "VK saveWallPhoto error"))
    photo = saved["response"][0]
    return f"photo{photo['owner_id']}_{photo['id']}"

def vk_upload_video_to_wall(access_token: str, group_id: str, video_data: bytes, filename: str = "video.mp4", title: str = "", description: str = "") -> str:
    clean_id = group_id.lstrip("-")
    r = http_requests.post(
        "https://api.vk.com/method/video.save",
        data={
            "group_id": clean_id,
            "name": title or filename,
            "description": description,
            "wallpost": 0,
            "access_token": access_token,
            "v": VK_API_VERSION,
        },
        timeout=15,
    )
    data = r.json()
    if "error" in data:
        raise ValueError(data["error"].get("error_msg", "VK video.save error"))
    resp = data["response"]
    upload_url = resp["upload_url"]
    owner_id = resp["owner_id"]
    video_id = resp["video_id"]

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "mp4"
    mime_map = {"mp4": "video/mp4", "mov": "video/quicktime", "webm": "video/webm", "avi": "video/x-msvideo", "mkv": "video/x-matroska"}
    content_type = mime_map.get(ext, "video/mp4")

    r2 = http_requests.post(upload_url, files={"video_file": (filename, video_data, content_type)}, timeout=600)
    r2.raise_for_status()
    upload_result = r2.json()
    if "error" in upload_result:
        err = upload_result["error"]
        raise ValueError(err if isinstance(err, str) else str(err))

    return f"video{owner_id}_{video_id}"

def vk_upload_doc_to_wall(access_token: str, group_id: str, doc_data: bytes, filename: str = "doc.pdf", title: str = "") -> str:
    clean_id = group_id.lstrip("-")
    r = http_requests.get(
        "https://api.vk.com/method/docs.getWallUploadServer",
        params={"group_id": clean_id, "access_token": access_token, "v": VK_API_VERSION},
        timeout=10,
    )
    data = r.json()
    if "error" in data:
        raise ValueError(data["error"].get("error_msg", "VK docs.getWallUploadServer error"))
    upload_url = data["response"]["upload_url"]

    r2 = http_requests.post(upload_url, files={"file": (filename, doc_data, "application/octet-stream")}, timeout=300)
    r2.raise_for_status()
    upload_result = r2.json()
    if "error" in upload_result:
        err = upload_result["error"]
        raise ValueError(err if isinstance(err, str) else str(err))
    if not upload_result.get("file"):
        raise ValueError(f"Неожиданный ответ от VK при загрузке документа: {upload_result}")

    r3 = http_requests.post(
        "https://api.vk.com/method/docs.save",
        data={
            "file": upload_result["file"],
            "title": title or filename,
            "access_token": access_token,
            "v": VK_API_VERSION,
        },
        timeout=15,
    )
    saved = r3.json()
    if "error" in saved:
        raise ValueError(saved["error"].get("error_msg", "VK docs.save error"))
    resp = saved["response"]
    doc = resp["doc"] if isinstance(resp, dict) and "doc" in resp else (resp[0] if isinstance(resp, list) else resp)
    return f"doc{doc['owner_id']}_{doc['id']}"

def vk_wall_post(access_token: str, group_id: str, message: str, attachments: List[str] = []) -> int:
    clean_id = group_id.lstrip("-")
    params: dict = {
        "owner_id": f"-{clean_id}",
        "from_group": 1,
        "message": message,
        "access_token": access_token,
        "v": VK_API_VERSION,
    }
    if attachments:
        params["attachments"] = ",".join(attachments)
    r = http_requests.post(
        "https://api.vk.com/method/wall.post",
        data=params,
        timeout=15,
    )
    data = r.json()
    if "error" in data:
        raise ValueError(data["error"].get("error_msg", "VK wall.post error"))
    return data["response"]["post_id"]

# ── Telegram API helpers ─────────────────────────────────────────────────────

TG_CAPTION_LIMIT = 1024
TG_MESSAGE_LIMIT = 4096

def tg_api(bot_token: str, method: str, **kwargs) -> dict:
    url = f"https://api.telegram.org/bot{bot_token}/{method}"
    r = http_requests.post(url, timeout=kwargs.pop("_timeout", 30), **kwargs)
    try:
        data = r.json()
    except Exception:
        raise ValueError(f"Некорректный ответ Telegram ({r.status_code})")
    if not data.get("ok"):
        raise ValueError(data.get("description", "Telegram API error"))
    return data["result"]

def tg_get_chat_title(bot_token: str, chat_id: str) -> str:
    res = tg_api(bot_token, "getChat", data={"chat_id": chat_id})
    return res.get("title") or res.get("username") or str(res.get("id", ""))

def _resolve_media_bytes(item: dict, backend_base: str) -> tuple[bytes, str]:
    fname = os.path.basename(item["url"])
    fpath = os.path.join(UPLOAD_DIR, fname)
    if os.path.exists(fpath):
        with open(fpath, "rb") as f:
            return f.read(), item.get("filename") or fname
    file_url = f"{backend_base}{item['url']}"
    resp = http_requests.get(file_url, timeout=120)
    resp.raise_for_status()
    return resp.content, item.get("filename") or fname

def tg_send_post(bot_token: str, chat_id: str, message: str, media: List[dict], backend_base: str) -> List[int]:
    photos_videos = [m for m in media if m.get("type") in ("image", "video")]
    docs = [m for m in media if m.get("type") == "doc"]
    posted_ids: List[int] = []

    caption_with_media = len(message) <= TG_CAPTION_LIMIT and (photos_videos or docs)
    sent_text_separately = False

    if message and not caption_with_media:
        for chunk_start in range(0, len(message), TG_MESSAGE_LIMIT):
            chunk = message[chunk_start:chunk_start + TG_MESSAGE_LIMIT]
            res = tg_api(bot_token, "sendMessage", data={"chat_id": chat_id, "text": chunk})
            posted_ids.append(res.get("message_id"))
        sent_text_separately = True

    if photos_videos:
        if len(photos_videos) == 1:
            item = photos_videos[0]
            file_bytes, fname = _resolve_media_bytes(item, backend_base)
            method = "sendPhoto" if item["type"] == "image" else "sendVideo"
            field = "photo" if item["type"] == "image" else "video"
            data = {"chat_id": chat_id}
            if not sent_text_separately and message:
                data["caption"] = message[:TG_CAPTION_LIMIT]
            res = tg_api(bot_token, method, data=data, files={field: (fname, file_bytes)}, _timeout=300)
            posted_ids.append(res.get("message_id"))
        else:
            files = {}
            media_payload = []
            for idx, item in enumerate(photos_videos[:10]):
                file_bytes, fname = _resolve_media_bytes(item, backend_base)
                attach_key = f"file{idx}"
                files[attach_key] = (fname, file_bytes)
                m = {"type": "photo" if item["type"] == "image" else "video", "media": f"attach://{attach_key}"}
                if idx == 0 and not sent_text_separately and message:
                    m["caption"] = message[:TG_CAPTION_LIMIT]
                media_payload.append(m)
            res = tg_api(
                bot_token, "sendMediaGroup",
                data={"chat_id": chat_id, "media": json.dumps(media_payload)},
                files=files, _timeout=600,
            )
            if isinstance(res, list):
                posted_ids.extend(m.get("message_id") for m in res)

    for idx, item in enumerate(docs):
        file_bytes, fname = _resolve_media_bytes(item, backend_base)
        data = {"chat_id": chat_id}
        if idx == 0 and not sent_text_separately and not photos_videos and message:
            data["caption"] = message[:TG_CAPTION_LIMIT]
        res = tg_api(bot_token, "sendDocument", data=data, files={"document": (fname, file_bytes)}, _timeout=300)
        posted_ids.append(res.get("message_id"))

    if not posted_ids and message:
        res = tg_api(bot_token, "sendMessage", data={"chat_id": chat_id, "text": message[:TG_MESSAGE_LIMIT]})
        posted_ids.append(res.get("message_id"))

    return posted_ids

# ── VK Settings ───────────────────────────────────────────────────────────────

@app.get("/api/settings/vk")
def get_vk_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, group_id, group_name, connected_at FROM vk_settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    if not row:
        return {"connected": False}
    d = dict(row)
    d["connected"] = True
    return d

@app.post("/api/settings/vk")
def save_vk_settings(body: VkSettingsSave):
    try:
        group_name = vk_get_group_name(body.access_token, body.group_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM vk_settings WHERE id=1")
    exists = c.fetchone()
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    if exists:
        c.execute(
            "UPDATE vk_settings SET group_id=%s, access_token=%s, group_name=%s, connected_at=%s WHERE id=1",
            (body.group_id.lstrip("-"), body.access_token, group_name, now),
        )
    else:
        c.execute(
            "INSERT INTO vk_settings (id, group_id, access_token, group_name, connected_at) VALUES (1, %s, %s, %s, %s)",
            (body.group_id.lstrip("-"), body.access_token, group_name, now),
        )
    conn.commit()
    conn.close()
    return {"connected": True, "group_id": body.group_id.lstrip("-"), "group_name": group_name, "connected_at": now}

@app.post("/api/vk/oauth-exchange")
def vk_oauth_exchange(body: VkOAuthExchange):
    r = http_requests.get(
        "https://oauth.vk.com/access_token",
        params={
            "client_id": body.app_id,
            "client_secret": body.app_secret,
            "redirect_uri": "https://oauth.vk.com/blank.html",
            "code": body.code,
        },
        timeout=10,
    )
    data = r.json()
    if "error" in data:
        raise HTTPException(400, data.get("error_description") or data.get("error", "OAuth ошибка"))
    access_token = data.get("access_token")
    if not access_token:
        raise HTTPException(400, "Токен не получен от VK")
    try:
        group_name = vk_get_group_name(access_token, body.group_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM vk_settings WHERE id=1")
    exists = c.fetchone()
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    clean_id = body.group_id.lstrip("-")
    if exists:
        c.execute(
            "UPDATE vk_settings SET group_id=%s, access_token=%s, group_name=%s, connected_at=%s WHERE id=1",
            (clean_id, access_token, group_name, now),
        )
    else:
        c.execute(
            "INSERT INTO vk_settings (id, group_id, access_token, group_name, connected_at) VALUES (1, %s, %s, %s, %s)",
            (clean_id, access_token, group_name, now),
        )
    conn.commit()
    conn.close()
    return {"connected": True, "group_id": clean_id, "group_name": group_name, "connected_at": now}

@app.delete("/api/settings/vk")
def delete_vk_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM vk_settings WHERE id=1")
    conn.commit()
    conn.close()
    return {"connected": False}

# ── Group-scoped VK Settings ────────────────────────────────────────────────

@app.get("/api/groups/{gid}/settings/vk")
def get_group_vk_settings(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    require_group_member(gid, user_id, conn)
    c.execute("SELECT group_id, group_name, connected_at FROM vk_settings WHERE workspace_id=%s", (gid,))
    row = c.fetchone()
    conn.close()
    if not row:
        return {"connected": False}
    d = dict(row)
    d["connected"] = True
    return d

@app.post("/api/groups/{gid}/settings/vk")
def save_group_vk_settings(gid: int, body: VkSettingsSave, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может изменять настройки")
    try:
        group_name = vk_get_group_name(body.access_token, body.group_id)
    except ValueError as e:
        conn.close()
        raise HTTPException(400, str(e))
    clean_id = body.group_id.lstrip("-")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    c.execute("SELECT workspace_id FROM vk_settings WHERE workspace_id=%s", (gid,))
    exists = c.fetchone()
    if exists:
        c.execute(
            "UPDATE vk_settings SET group_id=%s, access_token=%s, group_name=%s, connected_at=%s WHERE workspace_id=%s",
            (clean_id, body.access_token, group_name, now, gid),
        )
    else:
        c.execute(
            "INSERT INTO vk_settings (workspace_id, group_id, access_token, group_name, connected_at) VALUES (%s, %s, %s, %s, %s)",
            (gid, clean_id, body.access_token, group_name, now),
        )
    conn.commit()
    conn.close()
    return {"connected": True, "group_id": clean_id, "group_name": group_name, "connected_at": now}

@app.delete("/api/groups/{gid}/settings/vk")
def delete_group_vk_settings(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может удалять настройки")
    c.execute("DELETE FROM vk_settings WHERE workspace_id=%s", (gid,))
    conn.commit()
    conn.close()
    return {"connected": False}

# ── Telegram Settings ────────────────────────────────────────────────────────

@app.get("/api/settings/telegram")
def get_tg_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, chat_id, chat_title, connected_at FROM tg_settings WHERE id=1")
    row = c.fetchone()
    conn.close()
    if not row:
        return {"connected": False}
    d = dict(row)
    d["connected"] = True
    return d

@app.post("/api/settings/telegram")
def save_tg_settings(body: TgSettingsSave):
    try:
        chat_title = tg_get_chat_title(body.bot_token, body.chat_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM tg_settings WHERE id=1")
    exists = c.fetchone()
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    if exists:
        c.execute(
            "UPDATE tg_settings SET bot_token=%s, chat_id=%s, chat_title=%s, connected_at=%s WHERE id=1",
            (body.bot_token, body.chat_id, chat_title, now),
        )
    else:
        c.execute(
            "INSERT INTO tg_settings (id, bot_token, chat_id, chat_title, connected_at) VALUES (1, %s, %s, %s, %s)",
            (body.bot_token, body.chat_id, chat_title, now),
        )
    conn.commit()
    conn.close()
    return {"connected": True, "chat_id": body.chat_id, "chat_title": chat_title, "connected_at": now}

@app.delete("/api/settings/telegram")
def delete_tg_settings():
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM tg_settings WHERE id=1")
    conn.commit()
    conn.close()
    return {"connected": False}

# ── Group-scoped Telegram Settings ──────────────────────────────────────────

@app.get("/api/groups/{gid}/settings/telegram")
def get_group_tg_settings(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    require_group_member(gid, user_id, conn)
    c.execute("SELECT chat_id, chat_title, connected_at FROM tg_settings WHERE workspace_id=%s", (gid,))
    row = c.fetchone()
    conn.close()
    if not row:
        return {"connected": False}
    d = dict(row)
    d["connected"] = True
    return d

@app.post("/api/groups/{gid}/settings/telegram")
def save_group_tg_settings(gid: int, body: TgSettingsSave, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может изменять настройки")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    # resolve chat_title via Telegram API
    chat_title = body.chat_id
    try:
        import requests as _req
        r = _req.get(f"https://api.telegram.org/bot{body.bot_token}/getChat",
                     params={"chat_id": body.chat_id}, timeout=5)
        data = r.json()
        if data.get("ok"):
            chat_title = data["result"].get("title") or data["result"].get("username") or body.chat_id
    except Exception:
        pass
    c.execute("SELECT workspace_id FROM tg_settings WHERE workspace_id=%s", (gid,))
    exists = c.fetchone()
    if exists:
        c.execute(
            "UPDATE tg_settings SET bot_token=%s, chat_id=%s, chat_title=%s, connected_at=%s WHERE workspace_id=%s",
            (body.bot_token, body.chat_id, chat_title, now, gid),
        )
    else:
        c.execute(
            "INSERT INTO tg_settings (workspace_id, bot_token, chat_id, chat_title, connected_at) VALUES (%s, %s, %s, %s, %s)",
            (gid, body.bot_token, body.chat_id, chat_title, now),
        )
    conn.commit()
    conn.close()
    return {"connected": True, "chat_id": body.chat_id, "chat_title": chat_title, "connected_at": now}

@app.delete("/api/groups/{gid}/settings/telegram")
def delete_group_tg_settings(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может удалять настройки")
    c.execute("DELETE FROM tg_settings WHERE workspace_id=%s", (gid,))
    conn.commit()
    conn.close()
    return {"connected": False}

# ── Auth ─────────────────────────────────────────────────────────────────────

@app.post("/api/auth/login")
def login(req: LoginRequest):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE email=%s", (req.email,))
    user = c.fetchone()
    if not user:
        conn.close()
        raise HTTPException(401, "Пользователь не найден")
    ph = user.get("password_hash")
    if ph and not verify_password(req.password, ph):
        conn.close()
        raise HTTPException(401, "Неверный пароль")
    user_id = user["id"]
    token = create_token(user_id)
    c.execute("SELECT g.id, g.name, g.description, g.avatar, gm.role, g.created_at FROM groups g JOIN group_members gm ON g.id = gm.group_id WHERE gm.user_id=%s ORDER BY g.id", (user_id,))
    groups = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"user": row_to_dict(user), "token": token, "groups": groups}

@app.post("/api/auth/register")
def register(req: RegisterRequest):
    if not req.name.strip():
        raise HTTPException(400, "Введите имя")
    if not req.email.strip():
        raise HTTPException(400, "Введите email")
    if len(req.password) < 6:
        raise HTTPException(400, "Пароль должен содержать минимум 6 символов")
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE email=%s", (req.email.lower().strip(),))
    if c.fetchone():
        conn.close()
        raise HTTPException(409, "Пользователь с таким email уже существует")
    avatar = "".join(p[0].upper() for p in req.name.strip().split()[:2])
    c.execute(
        "INSERT INTO users (name, email, role, avatar, password_hash) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (req.name.strip(), req.email.lower().strip(), "editor", avatar, hash_password(req.password)),
    )
    uid = c.fetchone()["id"]
    c.execute("SELECT id FROM groups ORDER BY id ASC LIMIT 1")
    default_group = c.fetchone()
    if default_group:
        c.execute("INSERT INTO group_members (group_id, user_id, role) VALUES (%s, %s, 'editor') ON CONFLICT DO NOTHING", (default_group["id"], uid))
    conn.commit()
    c.execute("SELECT * FROM users WHERE id=%s", (uid,))
    user = c.fetchone()
    token = create_token(uid)
    c.execute("SELECT g.id, g.name, g.description, g.avatar, gm.role, g.created_at FROM groups g JOIN group_members gm ON g.id = gm.group_id WHERE gm.user_id=%s ORDER BY g.id", (uid,))
    groups = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"user": row_to_dict(user), "token": token, "groups": groups}

# ── Groups ───────────────────────────────────────────────────────────────────

@app.post("/api/groups")
def create_group(req: GroupCreate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO groups (name, description, created_by) VALUES (%s, %s, %s) RETURNING id", (req.name, req.description, user_id))
    gid = c.fetchone()["id"]
    c.execute("INSERT INTO group_members (group_id, user_id, role) VALUES (%s, %s, 'admin')", (gid, user_id))
    conn.commit()
    c.execute("SELECT id, name, description, avatar, created_by, created_at FROM groups WHERE id=%s", (gid,))
    group = c.fetchone()
    conn.close()
    result = dict(group)
    result["role"] = "admin"
    return result

@app.get("/api/groups")
def get_my_groups(user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT g.id, g.name, g.description, g.avatar, gm.role, g.created_at FROM groups g JOIN group_members gm ON g.id = gm.group_id WHERE gm.user_id=%s ORDER BY g.id", (user_id,))
    groups = [dict(r) for r in c.fetchall()]
    conn.close()
    return groups

@app.get("/api/groups/{gid}")
def get_group(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    c.execute("SELECT id, name, description, avatar, created_by, created_at FROM groups WHERE id=%s", (gid,))
    group = c.fetchone()
    conn.close()
    if not group:
        raise HTTPException(404, "Группа не найдена")
    result = dict(group)
    result["role"] = role
    return result

@app.put("/api/groups/{gid}")
def update_group(gid: int, req: GroupUpdate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может изменять параметры группы")
    updates = []
    params = []
    for field, value in req.dict(exclude_unset=True).items():
        updates.append(f"{field}=%s")
        params.append(value)
    if not updates:
        conn.close()
        return get_group(gid, user_id)
    params.append(gid)
    c.execute(f"UPDATE groups SET {', '.join(updates)} WHERE id=%s", params)
    conn.commit()
    c.execute("SELECT id, name, description, avatar, created_by, created_at FROM groups WHERE id=%s", (gid,))
    group = c.fetchone()
    conn.close()
    result = dict(group)
    result["role"] = role
    return result

@app.delete("/api/groups/{gid}")
def delete_group(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может удалять группу")
    c.execute("DELETE FROM groups WHERE id=%s", (gid,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/groups/{gid}/members")
def get_group_members(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    require_group_member(gid, user_id, conn)
    c.execute("SELECT u.id, u.name, u.email, u.avatar, gm.role, gm.joined_at FROM users u JOIN group_members gm ON u.id = gm.user_id WHERE gm.group_id=%s ORDER BY u.id", (gid,))
    members = [dict(r) for r in c.fetchall()]
    conn.close()
    return members

@app.put("/api/groups/{gid}/members/{uid}/role")
def update_member_role(gid: int, uid: int, req: GroupMemberRoleUpdate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может изменять роли")
    c.execute("UPDATE group_members SET role=%s WHERE group_id=%s AND user_id=%s", (req.role, gid, uid))
    conn.commit()
    c.execute("SELECT u.id, u.name, u.email, u.avatar, gm.role, gm.joined_at FROM users u JOIN group_members gm ON u.id = gm.user_id WHERE gm.group_id=%s AND u.id=%s", (gid, uid))
    member = c.fetchone()
    conn.close()
    return dict(member) if member else None

@app.delete("/api/groups/{gid}/members/{uid}")
def remove_group_member(gid: int, uid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin" and user_id != uid:
        conn.close()
        raise HTTPException(403, "Вы можете удалить только себя из группы")
    c.execute("DELETE FROM group_members WHERE group_id=%s AND user_id=%s", (gid, uid))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/groups/{gid}/invites")
def create_invite_link(gid: int, req: InviteLinkCreate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может создавать ссылки приглашения")
    token = uuid.uuid4().hex
    expires_at = (datetime.utcnow() + timedelta(hours=req.expires_hours)).isoformat() + "Z"
    c.execute(
        "INSERT INTO invite_links (group_id, token, role, created_by, expires_at, max_uses) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id",
        (gid, token, req.role, user_id, expires_at, req.max_uses)
    )
    link_id = c.fetchone()["id"]
    conn.commit()
    c.execute("SELECT id, token, role, expires_at, used_count, max_uses, created_at FROM invite_links WHERE id=%s", (link_id,))
    link = c.fetchone()
    conn.close()
    return dict(link)

@app.get("/api/groups/{gid}/invites")
def get_invite_links(gid: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может просматривать ссылки приглашения")
    c.execute("SELECT id, token, role, expires_at, used_count, max_uses, created_at FROM invite_links WHERE group_id=%s ORDER BY id DESC", (gid,))
    links = [dict(r) for r in c.fetchall()]
    conn.close()
    return links

@app.delete("/api/groups/{gid}/invites/{link_id}")
def revoke_invite_link(gid: int, link_id: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role != "admin":
        conn.close()
        raise HTTPException(403, "Только администратор может отзывать ссылки")
    c.execute("DELETE FROM invite_links WHERE id=%s AND group_id=%s", (link_id, gid))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/invites/{token}")
def get_invite_preview(token: str):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT il.group_id, g.name as group_name, g.description as group_description, il.role, il.expires_at FROM invite_links il JOIN groups g ON il.group_id = g.id WHERE il.token=%s", (token,))
    link = c.fetchone()
    conn.close()
    if not link:
        raise HTTPException(404, "Ссылка приглашения не найдена")
    result = dict(link)
    result["group_id"] = link["group_id"]
    result["group_name"] = link["group_name"]
    result["group_description"] = link["group_description"]
    result["role"] = link["role"]
    result["expires_at"] = link["expires_at"]
    return result

@app.post("/api/invites/{token}/accept")
def accept_invite(token: str, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT group_id, role, expires_at, max_uses, used_count FROM invite_links WHERE token=%s", (token,))
    link = c.fetchone()
    conn.close()
    if not link:
        raise HTTPException(404, "Ссылка приглашения не найдена")
    if datetime.fromisoformat(link["expires_at"].rstrip("Z")) < datetime.utcnow():
        raise HTTPException(410, "Ссылка приглашения истекла")
    if link["max_uses"] and link["used_count"] >= link["max_uses"]:
        raise HTTPException(410, "Лимит использований ссылки исчерпан")
    gid = link["group_id"]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT role FROM group_members WHERE group_id=%s AND user_id=%s", (gid, user_id))
    if c.fetchone():
        conn.close()
        raise HTTPException(409, "Вы уже участник этой группы")
    c.execute("INSERT INTO group_members (group_id, user_id, role) VALUES (%s, %s, %s)", (gid, link["role"], user_id))
    c.execute("UPDATE invite_links SET used_count=used_count+1 WHERE token=%s", (token,))
    conn.commit()
    c.execute("SELECT g.id, g.name, g.description, g.avatar, gm.role, g.created_at FROM groups g JOIN group_members gm ON g.id = gm.group_id WHERE gm.user_id=%s AND g.id=%s", (user_id, gid))
    group = c.fetchone()
    conn.close()
    return dict(group) if group else {"ok": True}

# ── Users ────────────────────────────────────────────────────────────────────

@app.get("/api/users")
def get_users():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users ORDER BY id")
    rows = c.fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

@app.put("/api/users/{user_id}/role")
def update_role(user_id: int, body: UserUpdate):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET role=%s WHERE id=%s", (body.role, user_id))
    conn.commit()
    c.execute("SELECT * FROM users WHERE id=%s", (user_id,))
    user = c.fetchone()
    conn.close()
    return row_to_dict(user)

@app.post("/api/users")
def create_user(body: UserCreate):
    if not body.name.strip():
        raise HTTPException(400, "Введите имя")
    if not body.email.strip():
        raise HTTPException(400, "Введите email")
    if len(body.password) < 6:
        raise HTTPException(400, "Пароль должен содержать минимум 6 символов")
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE email=%s", (body.email.lower().strip(),))
    if c.fetchone():
        conn.close()
        raise HTTPException(409, "Пользователь с таким email уже существует")
    avatar = "".join(p[0].upper() for p in body.name.strip().split()[:2])
    c.execute(
        "INSERT INTO users (name, email, role, avatar, password_hash) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (body.name.strip(), body.email.lower().strip(), body.role, avatar, hash_password(body.password)),
    )
    uid = c.fetchone()["id"]
    conn.commit()
    c.execute("SELECT * FROM users WHERE id=%s", (uid,))
    user = c.fetchone()
    conn.close()
    return row_to_dict(user)

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE id=%s", (user_id,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404, "Пользователь не найден")
    c.execute("DELETE FROM users WHERE id=%s", (user_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Group-scoped Posts ──────────────────────────────────────────────────────

@app.get("/api/groups/{gid}/posts")
def get_group_posts(gid: int, status: Optional[str] = None, platform: Optional[str] = None, tag: Optional[str] = None, limit: int = 100, offset: int = 0, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    require_group_member(gid, user_id, conn)
    q = "SELECT p.*, u.name as author_name FROM posts p LEFT JOIN users u ON p.author_id=u.id WHERE p.group_id=%s"
    params: list = [gid]
    if status:
        q += " AND p.status=%s"; params.append(status)
    if platform:
        q += " AND p.platforms LIKE %s"; params.append(f'%"{platform}"%')
    if tag:
        q += " AND p.tags LIKE %s"; params.append(f'%"{tag}"%')
    q += " ORDER BY p.created_at DESC LIMIT %s OFFSET %s"
    params += [limit, offset]
    c.execute(q, params)
    rows = c.fetchall()
    c.execute("SELECT COUNT(*) FROM posts WHERE group_id=%s", (gid,))
    total = c.fetchone()["count"]
    conn.close()
    return {"posts": [row_to_dict(r) for r in rows], "total": total}

@app.post("/api/groups/{gid}/posts")
def create_group_post(gid: int, body: PostCreate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role == "observer":
        conn.close()
        raise HTTPException(403, "Наблюдатели не могут создавать посты")
    c.execute(
        "INSERT INTO posts (title,content,status,platforms,tags,scheduled_at,location_address,location_lat,location_lng,author_id,template_type,media,group_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (body.title, body.content, body.status, json.dumps(body.platforms), json.dumps(body.tags), body.scheduled_at, body.location_address, body.location_lat, body.location_lng, user_id, body.template_type, json.dumps([m.dict() for m in body.media]), gid),
    )
    pid = c.fetchone()["id"]
    conn.commit()
    c.execute("SELECT * FROM posts WHERE id=%s", (pid,))
    row = c.fetchone()
    conn.close()
    return row_to_dict(row)

@app.get("/api/groups/{gid}/posts/{post_id}")
def get_group_post(gid: int, post_id: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    require_group_member(gid, user_id, conn)
    c.execute("SELECT p.*, u.name as author_name FROM posts p LEFT JOIN users u ON p.author_id=u.id WHERE p.id=%s AND p.group_id=%s", (post_id, gid))
    row = c.fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Пост не найден")
    return row_to_dict(row)

@app.put("/api/groups/{gid}/posts/{post_id}")
def update_group_post(gid: int, post_id: int, body: PostUpdate, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role == "observer":
        conn.close()
        raise HTTPException(403, "Наблюдатели не могут редактировать посты")
    c.execute("SELECT id FROM posts WHERE id=%s AND group_id=%s", (post_id, gid))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404, "Пост не найден")
    updates, params = [], []
    if body.title      is not None: updates.append("title=%s");        params.append(body.title)
    if body.content    is not None: updates.append("content=%s");      params.append(body.content)
    if body.status     is not None: updates.append("status=%s");       params.append(body.status)
    if body.platforms  is not None: updates.append("platforms=%s");    params.append(json.dumps(body.platforms))
    if body.tags       is not None: updates.append("tags=%s");         params.append(json.dumps(body.tags))
    if body.scheduled_at is not None: updates.append("scheduled_at=%s"); params.append(body.scheduled_at)
    if body.media      is not None: updates.append("media=%s");        params.append(json.dumps([m.dict() for m in body.media]))
    if body.location_address is not None: updates.append("location_address=%s"); params.append(body.location_address)
    if body.location_lat is not None: updates.append("location_lat=%s"); params.append(body.location_lat)
    if body.location_lng is not None: updates.append("location_lng=%s"); params.append(body.location_lng)
    if updates:
        params.append(post_id)
        c.execute(f"UPDATE posts SET {', '.join(updates)} WHERE id=%s", params)
        conn.commit()
    c.execute("SELECT * FROM posts WHERE id=%s", (post_id,))
    row = c.fetchone()
    conn.close()
    return row_to_dict(row)

@app.delete("/api/groups/{gid}/posts/{post_id}")
def delete_group_post(gid: int, post_id: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role == "observer":
        conn.close()
        raise HTTPException(403, "Наблюдатели не могут удалять посты")
    c.execute("SELECT id FROM posts WHERE id=%s AND group_id=%s", (post_id, gid))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404)
    c.execute("DELETE FROM posts WHERE id=%s", (post_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/groups/{gid}/posts/{post_id}/publish")
def publish_group_post(gid: int, post_id: int, user_id: int = Depends(get_current_user_id)):
    conn = get_db()
    c = conn.cursor()
    role = require_group_member(gid, user_id, conn)
    if role == "observer":
        conn.close()
        raise HTTPException(403, "Наблюдатели не могут публиковать посты")
    c.execute("SELECT * FROM posts WHERE id=%s AND group_id=%s", (post_id, gid))
    post = c.fetchone()
    if not post:
        conn.close()
        raise HTTPException(404, "Пост не найден")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    c.execute(
        "UPDATE posts SET status='published',published_at=%s,views=%s,reactions=%s,comments=%s,shares=%s WHERE id=%s",
        (now, random.randint(300, 1500), random.randint(20, 120), random.randint(5, 40), random.randint(3, 25), post_id),
    )
    conn.commit()
    post_dict = row_to_dict(post)
    vk_post_id = None
    vk_error = None
    photo_errors: list = []
    platforms = post_dict.get("platforms", [])
    if "vk" in platforms:
        c.execute("SELECT group_id, access_token FROM vk_settings WHERE group_id=%s", (gid,))
        vk = c.fetchone()
        if vk:
            try:
                message = f"{post_dict['title']}\n\n{post_dict['content']}"
                attachments = []
                backend_base = os.getenv("BACKEND_URL", "https://backend-production-30d6.up.railway.app").rstrip("/")
                for media in post_dict.get("media", []):
                    if media.get("type") == "image":
                        try:
                            filename = media.get("filename", "image")
                            image_data = http_requests.get(f"{backend_base}/{media['url']}", timeout=15).content
                            photo_id = vk_upload_photo_to_wall(vk["access_token"], vk["group_id"], image_data, filename)
                            attachments.append(photo_id)
                        except Exception as e:
                            photo_errors.append(str(e))
                    elif media.get("type") == "video":
                        try:
                            filename = media.get("filename", "video.mp4")
                            video_data = http_requests.get(f"{backend_base}/{media['url']}", timeout=120).content
                            video_id = vk_upload_video_to_wall(vk["access_token"], vk["group_id"], video_data, filename, post_dict["title"])
                            attachments.append(video_id)
                        except Exception as e:
                            photo_errors.append(f"Video upload error: {str(e)}")
                r = http_requests.post(
                    "https://api.vk.com/method/wall.post",
                    data={
                        "owner_id": f"-{vk['group_id']}",
                        "message": message,
                        "attachments": ",".join(attachments),
                        "access_token": vk["access_token"],
                        "v": VK_API_VERSION,
                    },
                    timeout=30,
                )
                vk_result = r.json()
                if "response" in vk_result:
                    vk_post_id = vk_result["response"].get("post_id")
                else:
                    vk_error = vk_result.get("error", {}).get("error_msg", "Unknown VK error")
            except Exception as e:
                vk_error = str(e)
    c.execute("SELECT * FROM posts WHERE id=%s", (post_id,))
    row = c.fetchone()
    conn.close()
    result = row_to_dict(row)
    result["vk_post_id"] = vk_post_id
    result["vk_error"] = vk_error
    result["photo_errors"] = photo_errors
    return result

# ── Posts ────────────────────────────────────────────────────────────────────

@app.get("/api/posts")
def get_posts(status: Optional[str] = None, platform: Optional[str] = None, tag: Optional[str] = None, limit: int = 100, offset: int = 0):
    conn = get_db()
    c = conn.cursor()
    q = "SELECT p.*, u.name as author_name FROM posts p LEFT JOIN users u ON p.author_id=u.id WHERE 1=1"
    params: list = []
    if status:
        q += " AND p.status=%s";   params.append(status)
    if platform:
        q += " AND p.platforms LIKE %s"; params.append(f'%"{platform}"%')
    if tag:
        q += " AND p.tags LIKE %s";      params.append(f'%"{tag}"%')
    q += " ORDER BY p.created_at DESC LIMIT %s OFFSET %s"
    params += [limit, offset]
    c.execute(q, params)
    rows = c.fetchall()
    c.execute("SELECT COUNT(*) FROM posts")
    total = c.fetchone()["count"]
    conn.close()
    return {"posts": [row_to_dict(r) for r in rows], "total": total}

@app.get("/api/posts/{post_id}")
def get_post(post_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT p.*, u.name as author_name FROM posts p LEFT JOIN users u ON p.author_id=u.id WHERE p.id=%s", (post_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        raise HTTPException(404, "Пост не найден")
    return row_to_dict(row)

@app.post("/api/posts")
def create_post(body: PostCreate):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO posts (title,content,status,platforms,tags,scheduled_at,location_address,location_lat,location_lng,author_id,template_type,media) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (body.title, body.content, body.status, json.dumps(body.platforms), json.dumps(body.tags), body.scheduled_at, body.location_address, body.location_lat, body.location_lng, body.author_id, body.template_type, json.dumps([m.dict() for m in body.media])),
    )
    pid = c.fetchone()["id"]
    conn.commit()
    c.execute("SELECT * FROM posts WHERE id=%s", (pid,))
    row = c.fetchone()
    conn.close()
    return row_to_dict(row)

@app.put("/api/posts/{post_id}")
def update_post(post_id: int, body: PostUpdate):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM posts WHERE id=%s", (post_id,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404, "Пост не найден")
    updates, params = [], []
    if body.title      is not None: updates.append("title=%s");        params.append(body.title)
    if body.content    is not None: updates.append("content=%s");      params.append(body.content)
    if body.status     is not None: updates.append("status=%s");       params.append(body.status)
    if body.platforms  is not None: updates.append("platforms=%s");    params.append(json.dumps(body.platforms))
    if body.tags       is not None: updates.append("tags=%s");         params.append(json.dumps(body.tags))
    if body.scheduled_at is not None: updates.append("scheduled_at=%s"); params.append(body.scheduled_at)
    if body.media      is not None: updates.append("media=%s");        params.append(json.dumps([m.dict() for m in body.media]))
    if body.location_address is not None: updates.append("location_address=%s"); params.append(body.location_address)
    if body.location_lat is not None: updates.append("location_lat=%s"); params.append(body.location_lat)
    if body.location_lng is not None: updates.append("location_lng=%s"); params.append(body.location_lng)
    if updates:
        params.append(post_id)
        c.execute(f"UPDATE posts SET {', '.join(updates)} WHERE id=%s", params)
        conn.commit()
    c.execute("SELECT * FROM posts WHERE id=%s", (post_id,))
    row = c.fetchone()
    conn.close()
    return row_to_dict(row)

@app.delete("/api/posts/{post_id}")
def delete_post(post_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM posts WHERE id=%s", (post_id,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404)
    c.execute("DELETE FROM posts WHERE id=%s", (post_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/api/posts/{post_id}/publish")
def publish_post(post_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM posts WHERE id=%s", (post_id,))
    post = c.fetchone()
    if not post:
        conn.close()
        raise HTTPException(404)
    now = datetime.now().strftime("%Y-%m-%dT%H:%M")
    c.execute(
        "UPDATE posts SET status='published',published_at=%s,views=%s,reactions=%s,comments=%s,shares=%s WHERE id=%s",
        (now, random.randint(300, 1500), random.randint(20, 120), random.randint(5, 40), random.randint(3, 25), post_id),
    )
    conn.commit()

    post_dict = row_to_dict(post)
    vk_post_id = None
    vk_error = None
    photo_errors: list = []

    platforms = post_dict.get("platforms", [])
    if "vk" in platforms:
        c.execute("SELECT group_id, access_token FROM vk_settings WHERE id=1")
        vk = c.fetchone()
        if vk:
            try:
                message = f"{post_dict['title']}\n\n{post_dict['content']}"
                attachments = []
                backend_base = os.getenv("BACKEND_URL", "https://backend-production-30d6.up.railway.app").rstrip("/")
                for item in (post_dict.get("media") or []):
                    item_type = item.get("type")
                    if item_type not in ("image", "video", "doc"):
                        continue
                    fname = os.path.basename(item["url"])
                    orig_name = item.get("filename") or fname
                    fpath = os.path.join(UPLOAD_DIR, fname)
                    try:
                        if os.path.exists(fpath):
                            with open(fpath, "rb") as f:
                                file_data = f.read()
                        else:
                            file_url = f"{backend_base}{item['url']}"
                            resp = http_requests.get(file_url, timeout=120)
                            resp.raise_for_status()
                            file_data = resp.content
                        if item_type == "image":
                            att = vk_upload_photo_to_wall(vk["access_token"], vk["group_id"], file_data, fname)
                        elif item_type == "video":
                            att = vk_upload_video_to_wall(
                                vk["access_token"], vk["group_id"], file_data, fname,
                                title=post_dict.get("title", ""),
                                description=post_dict.get("content", ""),
                            )
                        else:
                            att = vk_upload_doc_to_wall(
                                vk["access_token"], vk["group_id"], file_data, orig_name,
                                title=post_dict.get("title", "") or orig_name,
                            )
                        attachments.append(att)
                    except Exception as media_err:
                        msg = str(media_err)
                        if any(kw in msg.lower() for kw in [
                            "unavailable with group auth",
                            "group authorization",
                            "access denied",
                            "this action is not available",
                            "community token",
                            "group token",
                            "error_code: 15",
                            "no access to call this method",
                        ]):
                            msg = f"Нет прав на загрузку {item_type}. Получите пользовательский токен в Настройках (кнопка «Получить токен ВК»)"
                        photo_errors.append(msg)
                vk_post_id = vk_wall_post(vk["access_token"], vk["group_id"], message, attachments)
                if photo_errors:
                    notif_msg = (
                        f"Пост «{post_dict['title']}» опубликован в ВКонтакте, "
                        f"но {len(photo_errors)} фото не загружено: {photo_errors[0]}"
                    )
                    notif_type = "warning"
                else:
                    notif_msg = f"Пост «{post_dict['title']}» опубликован в группу ВКонтакте"
                    notif_type = "success"
                c.execute(
                    "INSERT INTO notifications (user_id, message, type, is_read) VALUES (1, %s, %s, 0)",
                    (notif_msg, notif_type),
                )
                conn.commit()
            except Exception as e:
                vk_error = str(e)
                c.execute(
                    "INSERT INTO notifications (user_id, message, type, is_read) VALUES (1, %s, %s, 0)",
                    (f"Ошибка публикации в VK: {vk_error}", "error"),
                )
                conn.commit()

    tg_message_ids: list = []
    tg_error = None
    if "telegram" in platforms:
        c.execute("SELECT bot_token, chat_id FROM tg_settings WHERE id=1")
        tg = c.fetchone()
        if tg:
            try:
                tg_message = f"{post_dict['title']}\n\n{post_dict['content']}" if post_dict.get("title") else post_dict.get("content", "")
                backend_base = os.getenv("BACKEND_URL", "https://backend-production-30d6.up.railway.app").rstrip("/")
                tg_message_ids = tg_send_post(
                    tg["bot_token"], tg["chat_id"], tg_message,
                    post_dict.get("media") or [], backend_base,
                )
                c.execute(
                    "INSERT INTO notifications (user_id, message, type, is_read) VALUES (1, %s, %s, 0)",
                    (f"Пост «{post_dict['title']}» опубликован в Telegram", "success"),
                )
                conn.commit()
            except Exception as e:
                tg_error = str(e)
                c.execute(
                    "INSERT INTO notifications (user_id, message, type, is_read) VALUES (1, %s, %s, 0)",
                    (f"Ошибка публикации в Telegram: {tg_error}", "error"),
                )
                conn.commit()

    c.execute("SELECT * FROM posts WHERE id=%s", (post_id,))
    row = c.fetchone()
    conn.close()
    result = row_to_dict(row)
    if vk_post_id is not None:
        result["vk_post_id"] = vk_post_id
    if vk_error is not None:
        result["vk_error"] = vk_error
    if photo_errors:
        result["vk_photo_errors"] = photo_errors
    if tg_message_ids:
        result["tg_message_ids"] = tg_message_ids
    if tg_error is not None:
        result["tg_error"] = tg_error
    return result

# ── Calendar ─────────────────────────────────────────────────────────────────

@app.get("/api/calendar")
def get_calendar(start: str = Query(...), end: str = Query(...)):
    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT id,title,status,platforms,tags,scheduled_at,published_at,created_at FROM posts "
        "WHERE (scheduled_at BETWEEN %s AND %s) OR (published_at BETWEEN %s AND %s) OR (created_at BETWEEN %s AND %s) "
        "ORDER BY COALESCE(scheduled_at, published_at, created_at)",
        (start, end, start, end, start, end),
    )
    rows = c.fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]

# ── Templates ─────────────────────────────────────────────────────────────────

@app.get("/api/templates")
def get_templates():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM templates")
    rows = c.fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        try:
            d["fields"] = json.loads(d["fields"])
        except Exception:
            d["fields"] = []
        result.append(d)
    return result

@app.post("/api/generate-text")
def generate_text(body: GenerateRequest):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM templates WHERE type=%s", (body.template_type,))
    tmpl = c.fetchone()
    conn.close()
    if not tmpl:
        raise HTTPException(404, "Шаблон не найден")
    text = tmpl["template_text"]
    for k, v in body.fields.items():
        text = text.replace(f"{{{k}}}", v)
    titles = {
        "announcement": "Анонс: " + body.fields.get("event_name", ""),
        "results":      "Итоги: " + body.fields.get("event_name", ""),
        "vacancy":      "Вакансия: " + body.fields.get("position", ""),
        "grant":        body.fields.get("grant_name", ""),
    }
    return {"text": text, "title": titles.get(body.template_type, "Новый пост")}

# ── Analytics ─────────────────────────────────────────────────────────────────

@app.get("/api/analytics/summary")
def analytics_summary():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM posts");                         total     = c.fetchone()["count"]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='published'"); published = c.fetchone()["count"]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='scheduled'"); scheduled = c.fetchone()["count"]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='draft'");     drafts    = c.fetchone()["count"]
    c.execute("SELECT SUM(views) v,SUM(reactions) r,SUM(comments) c,SUM(shares) sh FROM posts WHERE status='published'")
    s = c.fetchone()
    c.execute(
        "SELECT id,title,views,reactions,comments,shares,published_at FROM posts "
        "WHERE status='published' ORDER BY (views+reactions*3+comments*2+shares*4) DESC LIMIT 5"
    )
    top = c.fetchall()
    pl_stats = []
    for pl in ["vk", "telegram"]:
        c.execute(
            "SELECT COUNT(*) cnt,SUM(views) v,SUM(reactions) r FROM posts "
            "WHERE platforms LIKE %s AND status='published'",
            (f'%"{pl}"%',),
        )
        ps = c.fetchone()
        pl_stats.append({"platform": pl, "count": ps["cnt"] or 0, "views": ps["v"] or 0, "reactions": ps["r"] or 0})
    conn.close()
    total_views = s["v"] or 0
    eng = round(((s["r"] or 0) + (s["c"] or 0)) / max(total_views, 1) * 100, 1)
    return {
        "total_posts": total, "published": published, "scheduled": scheduled, "drafts": drafts,
        "total_views": total_views, "total_reactions": s["r"] or 0,
        "total_comments": s["c"] or 0, "total_shares": s["sh"] or 0,
        "avg_views": round(total_views / max(published, 1)),
        "engagement_rate": eng,
        "top_posts": [dict(r) for r in top],
        "platform_stats": pl_stats,
    }

@app.get("/api/analytics/timeline")
def analytics_timeline(period: str = "month"):
    days = {"week": 7, "month": 30, "quarter": 90}.get(period, 30)
    now = datetime.now()
    result = []
    conn = get_db()
    c = conn.cursor()
    for i in range(days - 1, -1, -1):
        day = now - timedelta(days=i)
        ds = day.strftime("%Y-%m-%d")
        c.execute(
            "SELECT SUM(views) v, SUM(reactions) r, COUNT(*) p FROM posts WHERE published_at LIKE %s",
            (ds + "%",),
        )
        row = c.fetchone()
        result.append({"date": ds, "label": day.strftime("%d.%m"), "views": row["v"] or 0, "reactions": row["r"] or 0, "posts": row["p"] or 0})
    conn.close()
    return result

@app.get("/api/analytics/export")
def analytics_export(start_date: str = Query(...), end_date: str = Query(...)):
    now = datetime.now()
    try:
        dt_start = datetime.strptime(start_date, "%Y-%m-%d")
        dt_end   = datetime.strptime(end_date,   "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Формат дат: YYYY-MM-DD")

    conn = get_db()
    c = conn.cursor()

    # Summary data — только за выбранный период
    date_filter = "published_at >= %s AND published_at < %s AND status='published'"
    end_next    = (dt_end + timedelta(days=1)).strftime("%Y-%m-%d")
    start_str   = dt_start.strftime("%Y-%m-%d")

    c.execute(f"SELECT COUNT(*) FROM posts WHERE {date_filter}", (start_str, end_next))
    published = c.fetchone()["count"]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='scheduled'"); scheduled = c.fetchone()["count"]
    c.execute("SELECT COUNT(*) FROM posts WHERE status='draft'");     drafts    = c.fetchone()["count"]
    c.execute(f"SELECT COUNT(*) FROM posts WHERE published_at >= %s AND published_at < %s", (start_str, end_next))
    total = c.fetchone()["count"]
    c.execute(
        f"SELECT SUM(views) v,SUM(reactions) r,SUM(comments) cm,SUM(shares) sh FROM posts WHERE {date_filter}",
        (start_str, end_next),
    )
    s = c.fetchone()
    total_views = s["v"] or 0
    eng = round(((s["r"] or 0) + (s["cm"] or 0)) / max(total_views, 1) * 100, 1)

    # Timeline: каждый день периода
    timeline = []
    delta = (dt_end - dt_start).days + 1
    for i in range(delta):
        day = dt_start + timedelta(days=i)
        ds  = day.strftime("%Y-%m-%d")
        c.execute("SELECT SUM(views) v, SUM(reactions) r, COUNT(*) p FROM posts WHERE published_at LIKE %s", (ds + "%",))
        row = c.fetchone()
        timeline.append({"date": ds, "label": day.strftime("%d.%m"), "views": row["v"] or 0, "reactions": row["r"] or 0, "posts": row["p"] or 0})

    # Platform stats — за период
    pl_stats = []
    for pl in ["vk", "telegram"]:
        c.execute(
            f"SELECT COUNT(*) cnt,SUM(views) v,SUM(reactions) r FROM posts WHERE platforms LIKE %s AND {date_filter}",
            (f'%"{pl}"%', start_str, end_next),
        )
        ps = c.fetchone()
        pl_stats.append({"platform": pl.upper(), "count": ps["cnt"] or 0, "views": ps["v"] or 0, "reactions": ps["r"] or 0})

    # Top posts — за период
    c.execute(
        f"SELECT title,views,reactions,comments,shares,published_at FROM posts WHERE {date_filter} "
        "ORDER BY (views+reactions*3+comments*2+shares*4) DESC LIMIT 10",
        (start_str, end_next),
    )
    top_posts = c.fetchall()
    conn.close()

    # ── Build Excel ──
    wb = openpyxl.Workbook()

    BLUE   = "1D4ED8"
    WHITE  = "FFFFFF"
    GRAY   = "F1F5F9"
    DARK   = "1E293B"

    header_font  = Font(bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", fgColor=BLUE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align   = Alignment(vertical="center")
    thin_border  = Border(
        bottom=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="E2E8F0"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border

    def style_data_row(ws, row, cols, shade=False):
        fill = PatternFill("solid", fgColor=GRAY) if shade else None
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_align
            cell.border    = thin_border
            if shade:
                cell.fill = fill

    period_label = f"{dt_start.strftime('%d.%m.%Y')} — {dt_end.strftime('%d.%m.%Y')}"

    # ── Лист 1: Сводка ──
    ws1 = wb.active
    ws1.title = "Сводка"
    ws1.row_dimensions[1].height = 30
    ws1["A1"] = f"Отчёт по аналитике — {period_label}"
    ws1["A1"].font = Font(bold=True, size=14, color=DARK)
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells("A1:B1")
    ws1.append([])

    headers = ["Показатель", "Значение"]
    ws1.append(headers)
    style_header_row(ws1, 3, 2)
    ws1.row_dimensions[3].height = 24

    rows = [
        ("Всего постов",       total),
        ("Опубликовано",       published),
        ("Запланировано",      scheduled),
        ("Черновиков",         drafts),
        ("Всего просмотров",   total_views),
        ("Реакции",            s["r"] or 0),
        ("Комментарии",        s["cm"] or 0),
        ("Репосты",            s["sh"] or 0),
        ("Средние просмотры",  round(total_views / max(published, 1))),
        ("Вовлечённость, %",   eng),
    ]
    for i, (label, val) in enumerate(rows, start=4):
        ws1.append([label, val])
        ws1.row_dimensions[i].height = 20
        style_data_row(ws1, i, 2, shade=(i % 2 == 0))

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 18

    # ── Лист 2: Динамика ──
    ws2 = wb.create_sheet("Динамика")
    ws2.append(["Дата", "Просмотры", "Реакции", "Публикаций"])
    style_header_row(ws2, 1, 4)
    ws2.row_dimensions[1].height = 24
    for i, row in enumerate(timeline, start=2):
        ws2.append([row["date"], row["views"], row["reactions"], row["posts"]])
        ws2.row_dimensions[i].height = 18
        style_data_row(ws2, i, 4, shade=(i % 2 == 0))
    for col, w in zip("ABCD", [14, 14, 12, 14]):
        ws2.column_dimensions[col].width = w

    # ── Лист 3: Площадки ──
    ws3 = wb.create_sheet("Площадки")
    ws3.append(["Площадка", "Публикаций", "Просмотры", "Реакции"])
    style_header_row(ws3, 1, 4)
    ws3.row_dimensions[1].height = 24
    for i, pl in enumerate(pl_stats, start=2):
        ws3.append([pl["platform"], pl["count"], pl["views"], pl["reactions"]])
        ws3.row_dimensions[i].height = 20
        style_data_row(ws3, i, 4, shade=(i % 2 == 0))
    for col, w in zip("ABCD", [16, 14, 14, 12]):
        ws3.column_dimensions[col].width = w

    # ── Лист 4: Топ постов ──
    ws4 = wb.create_sheet("Топ постов")
    ws4.append(["#", "Заголовок", "Просмотры", "Реакции", "Комментарии", "Репосты", "Дата публикации"])
    style_header_row(ws4, 1, 7)
    ws4.row_dimensions[1].height = 24
    for i, p in enumerate(top_posts, start=2):
        ws4.append([i - 1, p["title"], p["views"] or 0, p["reactions"] or 0, p["comments"] or 0, p["shares"] or 0, str(p["published_at"] or "")])
        ws4.row_dimensions[i].height = 20
        style_data_row(ws4, i, 7, shade=(i % 2 == 0))
    for col, w in zip("ABCDEFG", [4, 40, 12, 10, 14, 10, 22]):
        ws4.column_dimensions[col].width = w

    # ── Stream response ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    month_ru = ["январь","февраль","март","апрель","май","июнь","июль","август","сентябрь","октябрь","ноябрь","декабрь"]
    fname = f"аналитика_{month_ru[dt_start.month-1]}_{dt_start.year}.xlsx"
    fname_encoded = quote(fname, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname_encoded}"},
    )

# ── Notifications ─────────────────────────────────────────────────────────────

@app.get("/api/notifications")
def get_notifications(user_id: int = 1):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 20", (user_id,))
    rows = c.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.put("/api/notifications/{notif_id}/read")
def mark_read(notif_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE notifications SET is_read=1 WHERE id=%s", (notif_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Startup ──────────────────────────────────────────────────────────────────

@app.on_event("startup")
def startup():
    try:
        init_db()
        print("✅  MediaHub API запущен!  →  http://localhost:8000")
    except Exception as e:
        print(f"❌  init_db() FAILED: {e}")
        raise

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
